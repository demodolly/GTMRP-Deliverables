#!/usr/bin/env python3
"""Reorganize user-filled BRD Requirements Register into correct columns."""

import re
from copy import deepcopy
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

CTT_BRIDGE_BOILERPLATE = (
    "Enables FY27 reporting parity between Workfront and non-Workfront teams during the hybrid "
    "operating model. CTT/Activity ID remains required for SFDC lead creation; CTT decommission "
    "is not currently planned in FY27."
)

HYBRID_LEGACY_IMPACT = (
    "Hybrid Phase 1 delivery: UTM and channel standardization progresses while legacy CCID/DTID "
    "remain mandatory for SFDC lead creation, MSP reporting, and Last Touch attribution until "
    "CTT decommission."
)

LEGACY_BARRIER_SNIPPET = "Continued use of CCID & DTID"

CTT_REASON_TAIL = re.compile(
    r"\s*[oO]\s*that we can have the same fields in CTT as we do in Workfront.*",
    re.DOTALL,
)


def clean_ctt_reason(text: str) -> str:
    if not text:
        return text
    cleaned = CTT_REASON_TAIL.sub("", text).strip()
    cleaned = re.sub(r"\s+", " ", cleaned)
    return cleaned.rstrip(" o").strip()


def split_reason_barrier_mixed(text: str):
    """Split mixed reason/barrier phrasing (e.g. UTM002)."""
    if not text:
        return text, None
    lower = text.lower()
    if "don't have a strategy" in lower or "no real direction" in lower:
        if " - but " in text:
            parts = text.split(" - but ", 1)
            return parts[0].strip(), parts[1].strip()
        if " but " in text:
            parts = text.split(" but ", 1)
            return parts[0].strip(), parts[1].strip()
    return text, None


def suggest_owner(row: dict) -> str:
    target = str(row.get("Target System") or row.get("Stakeholder / Platform") or "")
    rid = str(row.get("Requirement ID") or "")
    if "Stephen Watts" in str(row.get("Dependencies") or ""):
        return "Stephen Watts (AEM delivery)"
    if "Christy" in str(row.get("Barriers / Blockers") or ""):
        return "Christy Parker / Jim (Publishing process)"
    mapping = [
        ("Workfront", "Christy Parker (Workfront BCO)"),
        ("Tealium", "Konstantin Zobov (Tealium BCO)"),
        ("Eloqua", "Kim Braithwaite (Eloqua BCO)"),
        ("Marketo", "Devon Bell (Marketo BCO)"),
        ("Tray", "Jasper Peck (Tray BCO)"),
        ("SFDC", "Nadia Southerland (SFDC BCO)"),
        ("CTT", "CTT Dev Team / Amanda Pattenden"),
        ("OMS", "Data Engineering / OMS Team"),
        ("GTMPR", "Jessica Widener / Stephanie Boyd (GTMPR)"),
    ]
    for key, owner in mapping:
        if key.lower() in target.lower():
            return owner
    if rid.startswith("CTT") or rid.startswith("POR"):
        return "CTT Dev Team / GTMPR"
    if rid.startswith("DTID"):
        return "CTT Dev Team"
    return ""


def reorganize_row(row: dict) -> dict:
    out = deepcopy(row)
    reason = str(out.get("Reason Why Needed") or "").strip()
    barriers = str(out.get("Barriers / Blockers") or "").strip()
    impact = str(out.get("Impact on Main Project") or "").strip()
    comments = str(out.get("Comments") or "").strip()
    project = str(out.get("Project / BRD") or "")

    # CTT FY27: strip repeated bridge text from Reason → Impact
    if "CTT & CTT Request Portal" in project and CTT_REASON_TAIL.search(reason):
        out["Reason Why Needed"] = clean_ctt_reason(reason)
        if not impact:
            out["Impact on Main Project"] = CTT_BRIDGE_BOILERPLATE

    # DTID-001: long reason split
    if out.get("Requirement ID") == "DTID-001" and len(reason) > 120:
        out["Reason Why Needed"] = "Map legacy DTIDs to FY27 Reporting Channels without complex translation logic in reporting."
        if not impact:
            out["Impact on Main Project"] = (
                "Reporting teams can use FY27 channel definitions while DTIDs remain live for FY27; "
                "supports Workfront-first lookup with CTT fallback for non-Workfront activity."
            )

    # UTM002-style mixed reason/barrier
    new_reason, extra_barrier = split_reason_barrier_mixed(reason)
    if extra_barrier:
        out["Reason Why Needed"] = new_reason
        out["Barriers / Blockers"] = (
            f"{barriers}\n{extra_barrier}".strip() if barriers else extra_barrier
        )

    # Legacy CCID barrier pattern → populate Impact if empty
    if LEGACY_BARRIER_SNIPPET in str(out.get("Barriers / Blockers") or "") and not impact:
        out["Impact on Main Project"] = HYBRID_LEGACY_IMPACT

    # PUB001 cancelled: decision note to Comments
    if out.get("Requirement ID") == "PUB001" and out.get("Status") == "Cancelled":
        decision = (
            "Superseded by Content ID stitching approach — data layer Type/Sub-Type no longer required; "
            "values will be derived from Workfront Content ID to avoid Workfront vs form discrepancies."
        )
        out["Comments"] = decision if not comments else f"{comments}\n{decision}"

    # PUB002: "No historic updates" is a barrier not dependency
    if out.get("Requirement ID") == "PUB002":
        b = str(out.get("Barriers / Blockers") or "")
        if "No historic updates" in b:
            out["Comments"] = "Scope limited to new/updated pages; no retrospective page updates." 

    # Default blank status
    if not out.get("Status") or str(out.get("Status")).strip() == "":
        out["Status"] = "Not Started"

    # Suggest owner only where empty and status is actively tracked
    if not out.get("Owner / Responsible Team") and out.get("Status") not in (
        "Not Started",
        None,
        "",
    ):
        suggested = suggest_owner(out)
        if suggested:
            out["Owner / Responsible Team"] = suggested

    return out


def build_column_guide():
    return [
        ["Column", "What belongs here", "What does NOT belong here", "Example"],
        [
            "Reason Why Needed",
            "The business benefit — why this capability matters",
            "Blockers, dependencies, project impacts, status notes",
            "Align with FY27 Reporting Channels",
        ],
        [
            "Status",
            "Current delivery state (use dropdown)",
            "Long explanations — use Comments instead",
            "In Progress",
        ],
        [
            "Dependencies",
            "External inputs required BEFORE this can complete (people, systems, other reqs, dates)",
            "Why we're blocked — use Barriers for that",
            "GTMPR audit sign-off; Delivery from Stephen Watts scheduled 4 Sep",
        ],
        [
            "Barriers / Blockers",
            "What is stopping or complicating delivery RIGHT NOW",
            "Business benefits — those go in Reason Why Needed",
            "CCID still required for SFDC lead creation; Not all teams on Workfront",
        ],
        [
            "Impact on Main Project",
            "Effect on the wider programme if delayed, changed, or delivered (cross-project impact)",
            "Requirement-specific rationale",
            "Hybrid Phase 1: UTMs delivered while legacy CCID remains for SFDC",
        ],
        [
            "Owner / Responsible Team",
            "Named person/team accountable for delivery",
            "System names — use Target System for that",
            "Kim Braithwaite (Eloqua BCO)",
        ],
        [
            "Comments",
            "Decisions, scope notes, review feedback, anything else",
            "Core rationale that belongs in Reason/Barriers/Impact",
            "Cancelled — superseded by Content ID approach",
        ],
        [],
        ["Reference columns (from BRD — usually leave as-is)", "", "", ""],
        ["Project / BRD", "Which programme document this comes from", "", "UIF Phase 1"],
        ["Program / Initiative", "Section within the BRD (UTM, TEA, CTT Tool, etc.)", "", "TEA"],
        ["Stakeholder / Platform", "Who owns the capability in the BRD", "", "Tealium / Eloqua"],
        ["Target System", "Where the change is implemented", "", "Tealium / Eloqua"],
        ["Priority", "Must Have / Should Have / Nice to Have", "", "Must Have"],
        ["Phase / Timeline", "When it is planned", "", "Phase 1 / Q1 FY27"],
        ["BRD Document", "Source filename", "", ""],
    ]


def apply_formatting(ws, row_count, col_count):
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    ref_fill = PatternFill(start_color="E8EEF4", end_color="E8EEF4", fill_type="solid")
    track_fill = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ref_cols = set(range(1, 10)) | {15}  # A-I and O (BRD Document)
    track_cols = set(range(10, 15))  # J-N

    headers = [ws.cell(1, c).value for c in range(1, col_count + 1)]
    for c in range(1, col_count + 1):
        cell = ws.cell(1, c, headers[c - 1])
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    alt = PatternFill(start_color="F7F9FC", end_color="F7F9FC", fill_type="solid")
    for r in range(2, row_count + 1):
        for c in range(1, col_count + 1):
            cell = ws.cell(r, c)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            if r % 2 == 0:
                cell.fill = alt
            if c in track_cols:
                cell.fill = track_fill if r % 2 else PatternFill(
                    start_color="FFFDF5", end_color="FFFDF5", fill_type="solid"
                )

    widths = [28, 14, 52, 20, 22, 20, 12, 14, 36, 14, 28, 28, 32, 24, 40, 28]
    for i, w in enumerate(widths[:col_count], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "J2"  # freeze reference cols, scroll tracking cols
    ws.auto_filter.ref = f"A1:{get_column_letter(col_count)}{row_count}"


def main():
    src = "/tmp/user_register.xlsx"
    dst = "/workspace/BRD_Requirements_Register.xlsx"

    wb = load_workbook(src)
    ws = wb["Requirements Register"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    rows = []
    for r in range(2, ws.max_row + 1):
        row = {headers[c - 1]: ws.cell(r, c).value for c in range(1, len(headers) + 1)}
        if row.get("Requirement ID"):
            rows.append(reorganize_row(row))

    # Rewrite main sheet
    ws.delete_rows(2, ws.max_row)
    for i, row in enumerate(rows, 2):
        for c, h in enumerate(headers, 1):
            ws.cell(i, c, row.get(h))

    apply_formatting(ws, len(rows) + 1, len(headers))

    # Status dropdown from Lists sheet or create
    if "Lists" not in wb.sheetnames:
        wb.create_sheet("Lists")
    ls = wb["Lists"]
    statuses = [
        "Not Started",
        "In Progress",
        "Blocked",
        "In Review",
        "Complete",
        "Deferred",
        "Cancelled",
    ]
    ls.delete_rows(1, ls.max_row)
    ls.cell(1, 1, "Status")
    for i, s in enumerate(statuses, 2):
        ls.cell(i, 1, s)

    dv = DataValidation(type="list", formula1="=Lists!$A$2:$A$8", allow_blank=True)
    dv.error = "Please choose a status from the list"
    dv.prompt = "Select delivery status"
    ws.add_data_validation(dv)
    dv.add(f"J2:J{len(rows)+1}")

    # Column Guide
    if "Column Guide" in wb.sheetnames:
        del wb["Column Guide"]
    guide = wb.create_sheet("Column Guide", 0)
    for row in build_column_guide():
        guide.append(row)
    guide.column_dimensions["A"].width = 22
    guide.column_dimensions["B"].width = 42
    guide.column_dimensions["C"].width = 42
    guide.column_dimensions["D"].width = 36

    # Progress Summary
    if "Progress Summary" in wb.sheetnames:
        del wb["Progress Summary"]
    summary = wb.create_sheet("Progress Summary", 1)
    summary.append(["Project / BRD", "Total", "Not Started", "In Progress", "In Review", "Complete", "Deferred", "Cancelled", "Blocked"])
    from collections import defaultdict

    by_proj = defaultdict(list)
    for row in rows:
        by_proj[row["Project / BRD"]].append(row)
    for proj in sorted(by_proj):
        items = by_proj[proj]
        counts = {s: 0 for s in statuses}
        for it in items:
            st = str(it.get("Status") or "Not Started")
            counts[st] = counts.get(st, 0) + 1
        summary.append([
            proj,
            len(items),
            counts.get("Not Started", 0),
            counts.get("In Progress", 0),
            counts.get("In Review", 0),
            counts.get("Complete", 0),
            counts.get("Deferred", 0),
            counts.get("Cancelled", 0),
            counts.get("Blocked", 0),
        ])

    wb.save(dst)

    tracked = sum(1 for r in rows if r.get("Status") != "Not Started")
    print(f"Saved {len(rows)} requirements to {dst}")
    print(f"Tracked (status set): {tracked}")
    print(f"Remaining Not Started: {len(rows) - tracked}")


if __name__ == "__main__":
    main()
