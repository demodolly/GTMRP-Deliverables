#!/usr/bin/env python3
"""Generate BRD Requirements Register Excel from all workspace BRDs."""

from collections import Counter, defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADERS = [
    "Requirement ID",
    "Requirement Description",
    "Project / BRD",
    "Program / Initiative",
    "Stakeholder / Platform",
    "Target System",
    "Priority",
    "Phase / Timeline",
    "Reason Why Needed (from BRD)",
    "Status",
    "Dependencies",
    "Barriers / Blockers",
    "Impact on Main Project",
    "Owner / Responsible Team",
    "Related Requirement IDs",
    "BRD Document",
    "BRD Version / Date",
    "Comments",
]

requirements = []


def add(
    req_id,
    desc,
    project,
    program,
    stakeholder,
    target,
    priority,
    phase,
    reason,
    deps="",
    related="",
    doc="",
    version="",
    status="To be completed",
):
    requirements.append(
        {
            "Requirement ID": req_id,
            "Requirement Description": desc,
            "Project / BRD": project,
            "Program / Initiative": program,
            "Stakeholder / Platform": stakeholder,
            "Target System": target,
            "Priority": priority,
            "Phase / Timeline": phase,
            "Reason Why Needed (from BRD)": reason,
            "Status": status,
            "Dependencies": deps,
            "Barriers / Blockers": "",
            "Impact on Main Project": "",
            "Owner / Responsible Team": "",
            "Related Requirement IDs": related,
            "BRD Document": doc,
            "BRD Version / Date": version,
            "Comments": "",
        }
    )


def add_std(project, doc, version, rows):
    """rows: id, desc, program, target, priority, phase, reason"""
    for r in rows:
        add(r[0], r[1], project, r[2], r[3], r[3], r[4], r[5], r[6], doc=doc, version=version)


DOC1 = "Phase 1 - Unified Intelligence Framework - Requirements.docx"
PROG1 = "Unified Intelligence Framework (UIF) - Phase 1"

add_std(
    PROG1,
    DOC1,
    "v8.0 (27/02/2026)",
    [
        (
            "BAT001",
            "Execute End-to-End Business Acceptance Testing in a connected sandbox using mock payloads from Marketo and Eloqua to verify Tray.io logic processes attribution data into SFDC Lead Record per business rules before production deployment.",
            "BAT",
            "Tray.io / SFDC / Marketo / Eloqua",
            "Must Have",
            "Phase 1",
            "Validate end-to-end data flow before production deployment",
        ),
        (
            "UTM001",
            "All UTM values formatted in lowercase using hyphens as word separators and underscores as section separators for clean, consistent, queryable data.",
            "UTM",
            "Workfront / Stensul / Manual URL Builder",
            "Must Have",
            "Phase 1",
            "Maintain clean and queryable tracking data",
        ),
        (
            "UTM002",
            "Workfront supports optional parameters (utm_campaign, utm_term, utm_content, utm_creative, utm_eid) for granular campaign tracking beyond mandatory fields.",
            "UTM",
            "Workfront",
            "Should Have",
            "Phase 1",
            "Capture granular campaign details",
        ),
        (
            "UTM003",
            "URL parameters appended with ? before first parameter and & between subsequent parameters for syntactically correct tracking strings.",
            "UTM",
            "Workfront / Stensul / Manual URL Builder",
            "Must Have",
            "Phase 1",
            "Ensure tracking links function across browsers",
        ),
        (
            "UTM004",
            "Workfront automatically appends standardized URL parameters (utm_id, utm_medium, utm_source) plus legacy parameters (ccid, dtid) to all activation links.",
            "UTM",
            "Workfront",
            "Must Have",
            "Phase 1",
            "Consistent tracking across Marketo and Eloqua without breaking legacy reporting during MOPS transition",
        ),
        (
            "UTM005",
            "Manual URL Builder automatically appends standardized URL parameters (utm_id, utm_medium, utm_source) plus legacy parameters (ccid, dtid) to all activation links.",
            "UTM",
            "Manual URL Builder",
            "Must Have",
            "Phase 1",
            "Consistent ROI measurement for teams not on Workfront",
        ),
        (
            "UTM006",
            "Stensul logic automatically appends standardized URL parameters (utm_id, utm_medium, utm_source) plus legacy parameters (ccid, dtid) to all activation links.",
            "UTM",
            "Stensul",
            "Must Have",
            "Phase 1",
            "Consistent tracking across email activation tools",
        ),
        (
            "PUB001",
            "Publishing Team adds Content_Type and Content_Sub_Type metadata to the data layer of every webpage for content categorization and engagement analysis.",
            "PUB",
            "Website / Data Layer",
            "Must Have",
            "Phase 1",
            "Categorize and monitor page purpose and content engagement",
        ),
        (
            "PUB002",
            "Publishing Team adds placeholder for Content_Asset_ID to data layer of every webpage (empty until Workfront Standard Asset IDs available).",
            "PUB",
            "Website / Data Layer",
            "Must Have",
            "Phase 1 / Phase 2 ready",
            "Prepare website architecture for granular asset tracking without second sitewide update",
        ),
        (
            "TEA001",
            "Tealium captures and publishes UTM values to Eloqua upon form submission to preserve marketing attribution during lead capture.",
            "TEA",
            "Tealium / Eloqua",
            "Must Have",
            "Phase 1",
            "Preserve attribution data during lead capture",
        ),
        (
            "TEA002",
            "Tealium captures and persists Source Domain URL, Entry URL (with parameters), and CTA URL for session duration; writes alongside Conversion URL to Eloqua/Marketo hidden fields on form submission.",
            "TEA",
            "Tealium / Eloqua / Marketo",
            "Must Have",
            "Phase 1",
            "Complete map of customer journey at conversion",
        ),
        (
            "TEA003",
            "Tealium captures CP_GUTC, C_FPID and C_ECID and passes into hidden form fields during submission for persistent link between web session and lead record.",
            "TEA",
            "Tealium / Eloqua / Marketo",
            "Must Have",
            "Phase 1",
            "Establish persistent link for future touchpoint matching (Phase 2)",
        ),
        (
            "TEA004",
            "Tealium captures utm_id from URL and writes to both UTM_ID and CCID hidden fields on Eloqua form for duplicate values in Transaction CDO (Snowflake + legacy compatibility).",
            "TEA",
            "Tealium / Eloqua",
            "Must Have",
            "Phase 1",
            "Accurate Snowflake publishing while maintaining backward compatibility for legacy reports",
        ),
        (
            "TEA005",
            "Tealium maintains current 24-hour persistence logic for UTM values so attribution assigned even if user does not convert immediately.",
            "TEA",
            "Tealium",
            "Must Have",
            "Phase 1",
            "Retain attribution for delayed conversions",
        ),
        (
            "TEA006",
            "Remove CCID Page logic (impacts ~21% of records, many campaign-independent) to simplify tracking architecture.",
            "TEA",
            "Tealium / Eloqua",
            "Should Have",
            "Phase 1",
            "Simplify architecture without significant data integrity loss",
        ),
        (
            "TEA007",
            "Marketing Initiative ID and Drive to Channel fields set as First-Touch persistent so not overwritten by subsequent non-initiative interactions within same session.",
            "TEA",
            "Tealium / SFDC",
            "Must Have",
            "Phase 1",
            "Protect original attribution from overwrite",
        ),
        (
            "TEA008",
            "Remove logic that scrapes ccid value from page HTML when URL/Keyword search parameters absent to prevent incorrect alignment to static Content Offer.",
            "TEA",
            "Tealium",
            "Must Have",
            "Phase 1",
            "Prevent mucky data from false attribution",
        ),
        (
            "ELQ001",
            "All Eloqua forms include hidden fields for utm_term, utm_content, and utm_id for Tealium-captured attribution data into Transaction CDO.",
            "ELQ",
            "Eloqua",
            "Must Have",
            "Phase 1",
            "Pass Tealium attribution into Eloqua pipeline",
        ),
        (
            "ELQ002",
            "Tealium captures all 7 UTM parameters from Entry URL and writes to Eloqua hidden fields for Snowflake publication.",
            "ELQ",
            "Tealium / Eloqua",
            "Must Have",
            "Phase 1",
            "Standardized reporting format for Snowflake",
        ),
        (
            "ELQ003",
            "Amend creative HTML field name on Eloqua pages to map to utm_creative; map content HTML field to utm_content for correct pipeline extraction.",
            "ELQ",
            "Eloqua",
            "Must Have",
            "Phase 1",
            "Correct field mapping in data pipeline",
        ),
        (
            "ELQ004",
            "Create hidden field for Content Asset ID on Eloqua Landing Pages (no data until Phase 2) for structural Phase 2 readiness.",
            "ELQ",
            "Eloqua",
            "Must Have",
            "Phase 1 / Phase 2 ready",
            "Prepare platform for granular content tracking in future release",
        ),
        (
            "ELQ005",
            "Eloqua captures CP_GUTC, C_FPID and C_ECID during form submission for Phase 2 touchpoint matching.",
            "ELQ",
            "Eloqua",
            "Must Have",
            "Phase 1",
            "Link lead records to touchpoint traffic for Phase 2",
        ),
        (
            "ELQ006",
            "App Cloud Lite pulls all 7 UTM Parameters from Eloqua Transaction CDO to preserve original source tracking.",
            "ELQ",
            "App Cloud Lite",
            "Must Have",
            "Phase 1",
            "Preserve UTM data in integration pipeline",
        ),
        (
            "ELQ007",
            "App Cloud Lite passes all 7 UTM Parameters downstream to Tray.io.",
            "ELQ",
            "App Cloud Lite / Tray.io",
            "Must Have",
            "Phase 1",
            "Enable Tray.io to process standardized attribution",
        ),
        (
            "ELQ008",
            "App Cloud Lite continues passing CCID, DTID and OID downstream to Tray.io as Program ID, Drive To ID and Offer ID when UTM values not present.",
            "ELQ",
            "App Cloud Lite / Tray.io",
            "Must Have",
            "Phase 1",
            "Legacy fallback during transition",
        ),
        (
            "ELQ009",
            "Eloqua forms include hidden fields for Source Domain and Journey URLs for Tealium session data injection at conversion.",
            "ELQ",
            "Eloqua",
            "Must Have",
            "Phase 1",
            "Capture full session context",
        ),
        (
            "ELQ010",
            "Eloqua App Cloud Lite includes Source Domain and Journey URLs in outbound payload for integration layer processing.",
            "ELQ",
            "App Cloud Lite / Tray.io",
            "Must Have",
            "Phase 1",
            "Pass journey metadata downstream",
        ),
        (
            "ELQ011",
            "App Cloud Lite passes Content Asset ID downstream to Tray.io for future Workfront ID tracking.",
            "ELQ",
            "App Cloud Lite / Tray.io",
            "Must Have",
            "Phase 1 / Phase 2 ready",
            "Maintain content tracking when moving from Offer IDs to Workfront IDs",
        ),
        (
            "MKT001",
            "All Marketo forms include hidden fields for all 7 UTM parameters passed to Tray and Snowflake.",
            "MKT",
            "Marketo",
            "Must Have",
            "Phase 1",
            "Granular attribution capture on Marketo leads",
        ),
        (
            "MKT002",
            "All Marketo forms include hidden fields for Source Domain, Entry URL, CTA URL, and Conversion URL for full navigational path capture.",
            "MKT",
            "Marketo / Tray.io",
            "Must Have",
            "Phase 1",
            "Granular conversion reporting and journey pathing",
        ),
        (
            "MKT003",
            "All Marketo forms include hidden fields for CP_GUTC, C_FPID, C_ECID and Content Asset ID in datastream to Tray.io (Phase 1 placeholders).",
            "MKT",
            "Marketo / Tray.io",
            "Must Have",
            "Phase 1 / Phase 2 ready",
            "Structural preparation for Phase 2 transaction matching",
        ),
        (
            "TRY001",
            "Decouple SFDC Campaign association from legacy CCID/Program ID logic; Marketo passes native SFDC Campaign ID directly to Tray.io as standalone attribute.",
            "TRY",
            "Tray.io / SFDC",
            "Must Have",
            "Phase 1",
            "Content via Campaign membership while UTM_ID remains independent for channel attribution",
        ),
        (
            "TRY002",
            "Tray.io consumes utm_id from Marketo/Eloqua, uses as Program ID to lookup CCID in CTT tool, appends Program Name/Description before SFDC delivery.",
            "TRY",
            "Tray.io / CTT / SFDC",
            "Must Have",
            "Phase 1",
            "Consistent reporting on Driving Marketing Initiative",
        ),
        (
            "TRY003",
            "Tray.io Program ID mapping easily switchable from CTT IDs to Workfront IDs in future to minimize technical debt.",
            "TRY",
            "Tray.io",
            "Must Have",
            "Phase 1 / Phase 2",
            "Future-proof ID transition",
        ),
        (
            "TRY004",
            "Tray.io consumes all 7 UTM attributes from Eloqua (via App Cloud Lite) and Marketo; writes to Transaction Log and SFDC Lead Record.",
            "TRY",
            "Tray.io / SFDC",
            "Must Have",
            "Phase 1",
            "Standardized attribution visible for all marketing leads",
        ),
        (
            "TRY005",
            "Tray.io consumes Source Domain, Entry URL, CTA URL, Conversion URL from Marketo/Eloqua; writes to transaction log and SFDC.",
            "TRY",
            "Tray.io / SFDC",
            "Must Have",
            "Phase 1",
            "Report on customer journey paths and content performance",
        ),
        (
            "TRY006",
            "Tray continues consuming CTT parameters (CCID, DTID, OID) from App Cloud Lite during transition.",
            "TRY",
            "Tray.io",
            "Must Have",
            "Phase 1",
            "Maintain existing tracking while UTM standards implemented",
        ),
        (
            "TRY007",
            "Offline integrations and manual uploads continue utilizing legacy CCID, DTID, OID in Eloqua.",
            "TRY",
            "Tray.io / Eloqua",
            "Must Have",
            "Phase 1",
            "No disruption for channels not yet migrated",
        ),
        (
            "TRY008",
            "Tray.io maps OMS Vehicles to standard utm_medium and Sub Vehicles to utm_source via mapping table—or passes DTID if no match.",
            "TRY",
            "Tray.io / SFDC",
            "Must Have",
            "Phase 1",
            "Standardized CRM attribution with legacy fallback",
        ),
        (
            "TRY009",
            "Tray performs OMS lookup for Eloqua Offer IDs to append Offer name, description, and Offer type to Lead Record and Transaction Log.",
            "TRY",
            "Tray.io / SFDC",
            "Must Have",
            "Phase 1",
            "Granular visibility into offer types engaging leads",
        ),
        (
            "TRY010",
            "Individual page Content Asset ID captured in transaction log and sent to SFDC Lead Record (value populated from Phase 2).",
            "TRY",
            "Tray.io / SFDC",
            "Must Have",
            "Phase 1 / Phase 2 ready",
            "Granular content performance reporting beyond SFDC campaign level",
        ),
        (
            "TRY011",
            "Create placeholders in Tray.io Transaction Log for Content_Type and Content_Sub_Type for Phase 2 Workfront metadata.",
            "TRY",
            "Tray.io",
            "Must Have",
            "Phase 1 / Phase 2 ready",
            "Integration layer prepared for Workfront metadata",
        ),
        (
            "TRY012",
            "Tray.io applies waterfall logic when utm_id/ccid missing: Source Domain → Organic CCID; fallback Direct CCID.",
            "TRY",
            "Tray.io / SFDC",
            "Must Have",
            "Phase 1",
            "Uninterrupted lead creation with standardized attribution for non-campaign traffic",
        ),
        (
            "TRY013",
            "Maximum SLA of 15 minutes from form submission to SFDC lead creation.",
            "TRY",
            "Tray.io / SFDC",
            "Must Have",
            "Phase 1",
            "Prompt sales engagement while prospect interest is peak",
        ),
        (
            "TRY014",
            "Tray.io consumes Session_ID from Marketo/Eloqua to transaction log and SFDC (amended to first-party cookies in places).",
            "TRY",
            "Tray.io / SFDC",
            "Should Have",
            "Phase 1",
            "Persistent link for transaction-to-traffic matching",
        ),
        (
            "SFC001",
            "Create new SFDC Lead Record fields for all UTM parameters and CONTENT_ASSET_ID.",
            "SFC",
            "SFDC",
            "Must Have",
            "Phase 1",
            "CRM stores standardized attribution and content identifiers",
        ),
        (
            "SFC002",
            "Create placeholders on SFDC Lead Record for Content_Type and Content_Sub_Type for Phase 2 Workfront metadata.",
            "SFC",
            "SFDC",
            "Must Have",
            "Phase 1 / Phase 2 ready",
            "Store Workfront metadata when Phase 2 activated",
        ),
        (
            "SFC003",
            "SFDC Lead Record includes fields for Entry URL, CTA URL, and Conversion URL.",
            "SFC",
            "SFDC",
            "Must Have",
            "Phase 1",
            "Sales/marketing visibility of conversion trigger context",
        ),
        (
            "SFC004",
            "Placeholder field for Session_ID on SFDC Lead Record for Phase 2 deep-dive attribution.",
            "SFC",
            "SFDC",
            "Should Have",
            "Phase 1 / Phase 2 ready",
            "Facilitate Phase 2 touchpoint matching",
        ),
        (
            "SFC005",
            "SFDC VDC console displays Program ID from Tray.io instead of Activity ID tied to SFDC Campaign.",
            "SFC",
            "SFDC / VDC",
            "Must Have",
            "Phase 1",
            "Simplified accurate view of initiative driving revenue",
        ),
        (
            "REP001",
            "Capture URL of first page customer arrived at on .com at form submission as entry point of customer journey.",
            "REP",
            "Tealium / Reporting",
            "Must Have",
            "Phase 1",
            "Identify true entry point for journey analysis",
        ),
        (
            "REP002",
            "Capture URL of gated page and store as Conversion URL.",
            "REP",
            "Tealium / Reporting",
            "Must Have",
            "Phase 1",
            "Define specific converting offer page",
        ),
        (
            "REP003",
            "Capture URL of page when CTA was selected as CTA URL.",
            "REP",
            "Tealium / Reporting",
            "Must Have",
            "Phase 1",
            "Understand content context driving conversion",
        ),
        (
            "REP004",
            "Eloqua captures content as Offer ID; Marketo captures content in SFDC Campaign with Content Asset ID aligned to lead.",
            "REP",
            "Eloqua / Marketo / Tray.io",
            "Must Have",
            "Phase 1",
            "Track content asset performance across both platforms",
        ),
        (
            "REP005",
            "Use CTT Tool for content/campaign/program alignment when Eloqua Regular WebForm with Offer ID present.",
            "REP",
            "Tray.io / CTT",
            "Must Have",
            "Phase 1",
            "Enrich Eloqua WebForm leads with granular campaign data",
        ),
        (
            "REP006",
            "Check for DTID if utm_medium and utm_source missing during transition period.",
            "REP",
            "Tray.io / Reporting",
            "Must Have",
            "Phase 1",
            "Transition period fallback for channel attribution",
        ),
        (
            "REP007",
            "Append Content Type from CTT (Eloqua/Offer ID) or SFDC Campaign (Marketo).",
            "REP",
            "Tray.io / Reporting",
            "Must Have",
            "Phase 1",
            "Consistent content asset type visibility across platforms",
        ),
        (
            "REP008",
            "Set record as Organic Search if DTID and utm_medium missing but Entry URL populated.",
            "REP",
            "Tray.io / Reporting",
            "Must Have",
            "Phase 1",
            "Accurately credit SEO and unpaid search traffic",
        ),
        (
            "REP009",
            "Set record as Direct Search if DTID, utm_medium, and Referring URL all absent.",
            "REP",
            "Tray.io / Reporting",
            "Must Have",
            "Phase 1",
            "Catch-all classification for untrackable traffic",
        ),
        (
            "REP010",
            "Flag lead records failing attribution logic in Requires Review report on Ace Dashboard.",
            "REP",
            "Reporting / Ace Dashboard",
            "Must Have",
            "Phase 1",
            "Identify and fix new or broken traffic sources",
        ),
        (
            "REP011",
            "Update reporting filters to aggregate legacy DTIDs and new UTM parameters for hybrid view.",
            "REP",
            "Reporting",
            "Must Have",
            "FY27 refresh",
            "Comprehensive marketing performance view across hybrid ecosystem",
        ),
    ],
)

# Workfront CTT POC
DOC2 = "BRD - Workfront CTT Integration POC - Working Doc.docx"
PROG2 = "Workfront & CTT Integration POC"
for rid, desc, pri, reason in [
    ("BR-01", "Capture FY27 reporting data elements for ALL marketing activities regardless of Workfront or CTT origin.", "Must Have", "Core driver for FY27 measurement continuity"),
    ("BR-02", "Workfront teams obtain Activity/Offer IDs without manual ID administration.", "Must Have", "Remove adoption friction"),
    ("BR-03", "Workfront and non-Workfront activity reportable using same FY27 data elements.", "Must Have", "Consistent performance comparison"),
    ("BR-04", "No immediate SFDC or One Cisco CRM changes required.", "Must Have", "Protect OCC workstream"),
    ("BR-05", "POC produces evidence for scale, pause, or stop decision.", "Must Have", "Decision-quality investment"),
    ("BR-06", "POC documents CTT decommission blockers and dependencies.", "Must Have", "Align to decommission path"),
    ("BR-07", "Non-Workfront teams use CTT IDs with aligned future-state attributes.", "Should Have", "Support mixed adoption"),
    ("BR-08", "Interim solution defined as temporary with exit criteria.", "Must Have", "Prevent CTT extension perception"),
    ("BR-09", "Bridge data aligns with Adobe North Star model.", "Should Have", "Avoid rework"),
    ("BR-10", "Current CTT-dependent processes continue during POC.", "Must Have", "Business continuity"),
]:
    add(rid, desc, PROG2, "Business Requirements", "Cross-functional", "Workfront / CTT", pri, "POC", reason, doc=DOC2, version="v0.1 Draft (July 2026)")

for rid, desc, pri, reason in [
    ("FR-01", "Auto-create Activity ID in CTT when Workfront channel activation approved in pilot scope.", "Must Have", "Automate channel ID creation"),
    ("FR-02", "Auto-create Offer ID in CTT when Workfront content activation approved in pilot scope.", "Must Have", "Automate content ID creation"),
    ("FR-03", "Write back generated IDs to Workfront task/activation record.", "Must Have", "IDs available in Workfront"),
    ("FR-04", "Capture agreed FY27 attributes in both Workfront and CTT during bridge.", "Must Have", "Consistent FY27 data"),
    ("FR-05", "Maintain field mapping between Task ID, Activity ID, Offer ID, DTID, UTM_ID.", "Must Have", "Cross-system reporting alignment"),
    ("FR-06", "Support defined pilot activation type with known dependencies.", "Must Have", "Controlled POC scope"),
    ("FR-07", "Enable pilot dataset extraction with consistent FY27 attributes.", "Must Have", "Reporting proof"),
    ("FR-08", "Log integration events for audit and troubleshooting.", "Should Have", "Operational visibility"),
    ("FR-09", "CTT bridge fields configurable without SFDC changes.", "Must Have", "Bridge without CRM impact"),
    ("FR-10", "Governance controls for field ownership and change management.", "Should Have", "Sustainable operations"),
]:
    add(rid, desc, PROG2, "Functional Requirements", "Integration Team", "Workfront / CTT", pri, "POC Build", reason, related="BR-01 to BR-10", doc=DOC2, version="v0.1 Draft (July 2026)")

for rid, desc, reason in [
    ("NFR-01", "≥95% successful ID creation and write-back during validation.", "Prove integration reliability"),
    ("NFR-02", "≥95% completeness for required FY27 attributes on pilot activations.", "Data quality threshold"),
    ("NFR-03", "ID creation/write-back within agreed SLA after approval.", "Operational timeliness"),
    ("NFR-04", "Approved authentication; no hardcoded credentials.", "Secure integration"),
    ("NFR-05", "All ID/attribute changes traceable to source Workfront activation.", "Audit trail"),
    ("NFR-06", "Field mappings documented and version-controlled.", "Maintainability"),
    ("NFR-07", "No break to existing CTT-dependent processes during pilot.", "Business continuity"),
    ("NFR-08", "Architecture assessable for scale beyond pilot.", "Scale decision support"),
]:
    add(rid, desc, PROG2, "Non-Functional Requirements", "IT / Integration", "Workfront / CTT", "Must Have", "POC Validate", reason, doc=DOC2, version="v0.1 Draft (July 2026)")

for rid, desc, pri, reason in [
    ("RR-01", "Sample report/dashboard using FY27 data elements for Workfront and CTT pilot activity.", "Must Have", "Reporting continuity proof"),
    ("RR-02", "Segmentation by origin, channel vs content, and FY27 attributes.", "Must Have", "Flexible validation"),
    ("RR-03", "Compare pilot outputs to current attribution/funnel expectations.", "Must Have", "Validate alignment"),
    ("RR-04", "Validate consistent reporting for online, offline, non-Workfront activity.", "Must Have", "Hybrid reporting"),
    ("RR-05", "Document reporting gaps from pilot validation.", "Must Have", "Inform scale decision"),
    ("RR-06", "Identify reports consuming bridge data vs needing future changes.", "Should Have", "Decommission planning"),
]:
    add(rid, desc, PROG2, "Reporting Requirements", "Marketing Reporting", "Reporting / Snowflake", pri, "POC Validate", reason, doc=DOC2, version="v0.1 Draft (July 2026)")

# DTID Channel Alignment
DOC3 = "20260526-DTID Channel Alignment.docx"
PROG3 = "DTID Channel Alignment / Reporting Channel"
add_std(
    PROG3,
    DOC3,
    "v3.0 (26/05/2026)",
    [
        ("DTID-001", "Add REPORTING_CHANNEL field to MKW_OMS_IDS data table.", "Database", "CTT / OMS", "Must Have", "Q1 FY27", "Store future-state reporting channel mapping"),
        ("DTID-002", "Modify DTID Creation Screen with conditional dropdowns and read-only Reporting Channel.", "UI", "CTT Tool", "Must Have", "Q1 FY27", "Clean channel selection for new DTIDs"),
        ("DTID-003", "Modify DTID Edit Screen with conditional dropdowns and read-only Reporting Channel.", "UI", "CTT Tool", "Must Have", "Q1 FY27", "Edit existing DTIDs with reporting alignment"),
        ("DTID-004", "New DTID creation: restrict Channel dropdown to Table 1 LOV; hide deprecated channels.", "Logic", "CTT Tool", "Must Have", "Q1 FY27", "Prevent deprecated channel selection"),
        ("DTID-005", "New DTID creation: auto-populate Reporting Channel from mapping when Channel selected.", "Logic", "CTT Tool", "Must Have", "Q1 FY27", "Automate mapping at creation"),
        ("DTID-006", "Reporting Channel displays full English names (not lowercase/hyphenated).", "UI Standard", "CTT Tool", "Must Have", "Q1 FY27", "Human-readable categories"),
        ("DTID-007", "Edit mode: show saved REPORTING_CHANNEL on load; do not auto-remap.", "Logic", "CTT Tool", "Must Have", "Q1 FY27", "Preserve historical custom mappings"),
        ("DTID-008", "Edit mode: restrict channels to Table 2 Edit LOV.", "Logic", "CTT Tool", "Must Have", "Q1 FY27", "Controlled retro-mapping"),
        ("DTID-009", "Edit mode: mapping lookup only on active Channel dropdown change.", "Logic", "CTT Tool", "Must Have", "Q1 FY27", "Prevent accidental overwrite"),
        ("DTID-010", "Save validation: preserve REPORTING_CHANNEL if Channel unchanged.", "Logic", "CTT Tool", "Must Have", "Q1 FY27", "Data integrity on partial updates"),
        ("DTID-011", "Reporting Channel searchable/filterable/exportable in DTID search panel.", "Search", "CTT Tool", "Must Have", "Q1 FY27", "Audit and bulk analysis"),
        ("DTID-012", "Exclude Table 3 legacy codes from Edit dropdown while mapping in background.", "UI", "CTT Tool", "Must Have", "Q1 FY27", "Prevent retired code re-selection"),
        ("DTID-013", "One-time migration script for historical DTIDs using ID-level audit list.", "Data Migration", "CTT / Dev Team", "Must Have", "Pre-production", "Historical continuity for FY27 models"),
        ("DTID-014", "Expose Reporting Channel as view-only in front-end user portal.", "Search", "CTT Portal", "Should Have", "Q1 FY27", "User visibility of assignment"),
    ],
)
for r in requirements:
    if r["Requirement ID"] == "DTID-013":
        r["Dependencies"] = "Full Audit of DTIDs as of 09/06/2026"

# Phase 2, Decommission, Workfront Data, CTT FY27 - abbreviated loading via exec of remaining data in main
PROG4 = "Unified Intelligence Framework (UIF) - Phase 2"
DOC4 = "Unified Intelligence Framework - Phase 2 - Align Workfront IDs.docx"
add_std(PROG4, DOC4, "v1 (27/04/2026)", [
    ("INT001", "Workfront auto-generates Channel and Content IDs (Universal Key) for every activation.", "Intake", "Workfront", "Must Have", "Phase 2", "Eliminate manual ID creation"),
    ("INT002", "Workfront appends Channel ID to UTM_ID for all digital URLs.", "Intake", "Workfront", "Must Have", "Phase 2", "Hard-link conversions to initiatives"),
    ("INT003", "Intake provides Content ID to Publishing Team for page embedding.", "Intake", "Workfront / Publishing", "Must Have", "Phase 2", "Link conversions to content asset"),
    ("MAR001", "Capture Content ID from every AEM page in journey for attribution.", "MarTech", "Tealium / AEM", "Must Have", "Phase 2", "Attribute engagement to content assets"),
    ("MAR002", "Write last-touch Content ID into form submission in Marketo/Eloqua.", "MarTech", "Tealium / MAPs", "Must Have", "Phase 2", "Stamp lead with converting asset"),
    ("MAR003", "Investigate off-platform strategic data capture for future releases.", "MarTech", "MarTech Engineering", "Could Have", "Phase 2+", "Off-platform tracking gap analysis"),
    ("MAR004", "Identical utm_id and Content ID fields across Marketo and Eloqua.", "MarTech", "Marketo / Eloqua", "Must Have", "Phase 2", "Consistent Universal Key ingestion"),
    ("MAR005", "Identical legacy tag fields (ccid, dtid, oid) across both MAPs.", "MarTech", "Marketo / Eloqua", "Must Have", "Phase 2", "Parallel legacy reporting"),
    ("DSD001", "Tray differentiates Legacy vs Workfront IDs; queries correct source of truth.", "DSD", "Tray.io", "Must Have", "Phase 2", "Correct enrichment by ID type"),
    ("DSD002", "Tray stores Phase 2 metadata in transaction logs; no SFDC push during CRM freeze.", "DSD", "Tray.io", "Must Have", "Phase 2", "Validate flows without SFDC IT cost"),
    ("DSD003", "Synchronized staging ecosystem for Dual-Ingestion validation.", "DSD", "All platforms", "Must Have", "Phase 2", "Safe end-to-end testing"),
    ("CSI001", "Hybrid Reporting Mapping and Legacy Translation document for Snowflake.", "CSI", "CSI / Snowflake", "Must Have", "Phase 2", "YoY analysis without manual work"),
])

PROG5 = "CTT Decommission"
DOC5 = "20251125_Decommission Campaign Tagging & Tracking Tool.docx"
decom_rows = [
    ("IT001", "Remove CTT-Allocadia integration; use Workfront budget data.", "Integrated Tools", "Allocadia", "Must Have", "Decommission", "Align spend with Workfront"),
    ("IT002", "Decommission CTT picklist integration in Eloqua AOJs.", "Integrated Tools", "Eloqua", "Must Have", "Decommission", "Modern external data for AOJs"),
    ("IT003", "Remove CTT validation for SFDC lead creation; use Task ID/UTM sources.", "Integrated Tools", "SFDC / Tray.io", "Must Have", "Decommission", "Faster accurate lead creation"),
    ("IT004", "Remove CCID alignment from inbound call lead process.", "Integrated Tools", "SFDC", "Must Have", "Decommission", "Eliminate legacy in call flow"),
    ("IT005", "Align chat sessions to UTM_ID if marketing-driven.", "Integrated Tools", "SFDC", "Must Have", "Decommission", "Marketing-driven chat attribution"),
    ("IT006", "Remove Tray OCID extraction; use Task ID downstream.", "Integrated Tools", "Tray.io", "Must Have", "Decommission", "Standardize data exchange"),
    ("IT007", "Remove auto SFDC Campaign creation at CCID/OID levels.", "Integrated Tools", "Tray.io / SFDC", "Must Have", "Decommission", "Eliminate legacy campaign creation"),
    ("IT008", "Remove SFDC campaign member addition by CCID; use Task ID.", "Integrated Tools", "Tray.io / SFDC", "Must Have", "Decommission", "Accurate lead alignment"),
    ("IT009", "Decommission CTT-VDC Engage integration.", "Integrated Tools", "VDC Engage", "Must Have", "Decommission", "Simplify architecture"),
    ("MA001", "Amend Eloqua URLs to UTM values instead of CCID/DTID.", "Marketing Automation", "Eloqua", "Must Have", "Decommission", "Standardized outbound tracking"),
    ("MA002", "Audit high-use Eloqua forms; replace OID/CCID with Task ID.", "Marketing Automation", "Eloqua", "Must Have", "Decommission", "Data compliance on critical forms"),
    ("MA003", "Eloqua integrations remove OCID links; capture Task ID and UTMs.", "Marketing Automation", "Eloqua", "Must Have", "Decommission", "Eliminate legacy in integrations"),
    ("MA004", "Eloqua forms pass UTM values to App Cloud Lite, Tray, SFDC.", "Marketing Automation", "Eloqua / Tray.io", "Must Have", "Decommission", "Consistent modern tracking"),
    ("MA005", "Amend AOJs to UTM_ID instead of Activity ID.", "Marketing Automation", "Eloqua", "Must Have", "Decommission", "Granular Workfront-based reporting"),
    ("MA006", "Amend lead scoring to capture last UTM values not CTT values.", "Marketing Automation", "App Cloud Lite", "Must Have", "Decommission", "Current touchpoint-based scoring"),
    ("MA007", "Lead scoring prioritizes EID for events; remove CCID validation block.", "Marketing Automation", "App Cloud Lite", "Must Have", "Decommission", "Accurate event lead scoring"),
    ("MA008", "Amend Stensul URLs to UTM values not OCIDs.", "Marketing Automation", "Stensul", "Must Have", "Decommission", "Standardized email tracking"),
    ("MA009", "Manual uploads use Task ID not Activity ID.", "Marketing Automation", "Demand Intake Tool", "Must Have", "Decommission", "Standardized manual upload tracking"),
    ("MA010", "Event uploads validate EID not CCID.", "Marketing Automation", "Demand Intake Tool", "Must Have", "Decommission", "Standardized event identifier"),
    ("MA011", "Event enrichment uses Events data only.", "Marketing Automation", "Tray.io / Events", "Must Have", "Decommission", "Independent event enrichment"),
    ("MA012", "Remove DTID/OID from manual upload template and validations.", "Marketing Automation", "Demand Intake Tool", "Must Have", "Decommission", "Eliminate legacy from uploads"),
    ("MA013", "Capture utm_medium and utm_source on manual uploads.", "Marketing Automation", "Demand Intake Tool", "Must Have", "Decommission", "Channel analysis on offline uploads"),
    ("MA014", "MRF captures all reporting elements directly; no CTT/Workfront lookups.", "Marketing Automation", "MRF", "Must Have", "Decommission", "Self-contained data source"),
    ("MA015", "Event integrations use EID not CCID.", "Marketing Automation", "Event Integrations", "Must Have", "Decommission", "Standardized event data"),
    ("MA016", "Remove Tray OCID validations; use Workfront-based validations.", "Marketing Automation", "Tray.io", "Must Have", "Decommission", "Eliminate legacy in Tray"),
    ("MA017", "Hand-raiser event leads use EID and Events data only.", "Marketing Automation", "Tray.io", "Must Have", "Decommission", "Accurate event lead processing"),
    ("MA018", "Lean Data routing uses Task ID/UTM not OCID.", "Marketing Automation", "Lean Data", "Must Have", "Decommission", "Accurate routing post-CTT"),
    ("MA019", "Gong Engage/VDC sequences use Task ID not CCID/OID.", "Marketing Automation", "Gong Engage", "Must Have", "Decommission", "Correct engagement triggers"),
    ("MA020", "Stop CDF Eloqua pulls based on OCIDs.", "Marketing Automation", "CDF", "Must Have", "Decommission", "Decommission legacy pipeline"),
    ("MA021", "UTM Builders exclude CCID/DTID; use standardized UTMs.", "Marketing Automation", "Workfront / UTM Builder", "Must Have", "Decommission", "Enforce new methodology"),
    ("WI001", "Capture UTMs not OCIDs on driving website traffic.", "Web Integration", "Tealium / Website", "Must Have", "Decommission", "New standard in web analytics"),
    ("WI002", "Audit high-traffic pages for Task ID capture.", "Web Integration", "Website / AEM", "Must Have", "Decommission", "Journey attribution to tasks"),
    ("WI003", "Audit .com pages; remove all OCID URL references.", "Web Integration", "Website", "Must Have", "Decommission", "Privacy, security, URL standardization"),
    ("WI004", "Remove page logic referencing OCIDs; use Task ID.", "Web Integration", "Website / Tealium", "Must Have", "Decommission", "Legacy-free codebase"),
    ("RE001", "Reporting stops referencing CTT; uses Workfront data in Snowflake.", "Reporting", "Snowflake", "Must Have", "Decommission", "Governed live data only"),
    ("RE002", "Legacy-to-Task ID mapping table for historical aggregation.", "Reporting", "Snowflake", "Must Have", "Decommission", "YoY reporting"),
    ("RE003", "Reporting stops referencing CDF; uses channel source data.", "Reporting", "Snowflake", "Must Have", "Decommission", "Eliminate CDF dependency"),
    ("RE004", "Reports use Task ID/UTM filters only; remove OCID filters.", "Reporting", "Reporting", "Must Have", "Decommission", "Accurate new-attribute segmentation"),
    ("RE005", "New calculated fields from UTMs and Task ID.", "Reporting", "Reporting", "Must Have", "Decommission", "Future-proof reports"),
    ("RE006", "Retire reports relying solely on legacy CTT data.", "Reporting", "Reporting", "Must Have", "Decommission", "Guide users to new reports"),
    ("RE007", "Remove CDF Transaction Table references; report from source systems.", "Reporting", "CDF / Snowflake", "Must Have", "Decommission", "Simplify architecture"),
    ("ER001", "Capture UTM_ID and Page Task ID on .com arrival.", "Enriched Requirements", "Tealium / Website", "Should Have", "Decommission window", "Web reporting with dual IDs"),
    ("ER002", "Capture UTM_ID and Page Task ID in Eloqua on submission.", "Enriched Requirements", "Eloqua", "Should Have", "Decommission window", "Initiative + page context on lead"),
    ("ER003", "Pull UTM_ID for handraiser leads from Transaction CDO to Tray.", "Enriched Requirements", "App Cloud Lite", "Should Have", "Decommission window", "Downstream enrichment"),
    ("ER004", "Pull UTM_ID for scored leads from six CDOs to Tray.", "Enriched Requirements", "App Cloud Lite", "Should Have", "Decommission window", "Scored lead enrichment"),
    ("ER005", "Enrich lead from both UTM_ID and Page Task ID.", "Enriched Requirements", "Tray.io", "Should Have", "Decommission window", "Complete enriched lead"),
    ("ER006", "SFDC displays attributes from UTM_ID and Page Task ID.", "Enriched Requirements", "SFDC / VDC", "Should Have", "Decommission window", "Seller visibility"),
    ("ER007", "Reporting differentiates UTM_ID vs Page Task ID to prevent double-counting.", "Enriched Requirements", "Reporting", "Should Have", "Decommission window", "Accurate attribution"),
    ("ER008", "Capture both IDs when chat generates SFDC lead.", "Enriched Requirements", "SFDC / Chat", "Should Have", "Decommission window", "Chat lead marketing linkage"),
]
add_std(PROG5, DOC5, "v3.0 (08/12/2025)", decom_rows)

PROG6 = "Workfront Business-Ready Dataset (GTMRP)"
DOC6 = "20260513_Workfront Data Requirements - Attribution & Seller Enrichment.docx"
add_std(PROG6, DOC6, "v5.0 (21/05/2026)", [
    ("GTMRP001", "Business-ready Workfront dataset with clear terminology and minimal joins.", "GTMRP", "Snowflake", "Must Have", "Foundation", "Accelerate reporting"),
    ("GTMRP002", "Metadata repository and data dictionary for all delivered fields.", "GTMRP", "Data Governance", "Must Have", "Foundation", "Consistent metric logic"),
    ("GTMRP003", "Governance framework and intake process for new data elements.", "GTMRP", "Data Governance", "Must Have", "Ongoing", "Adapt while maintaining SSOT"),
    ("GTMRP004", "Recurring audit cycle for strategic data elements.", "GTMRP", "Data Governance", "Must Have", "Ongoing", "Keep dataset accurate"),
    ("GTMRP005", "Documentation and SOPs for self-service analysis.", "GTMRP", "GTMRP / Reporting", "Must Have", "Foundation", "Reduce support queries"),
    ("GEN001", "Aggregate strategic Workfront data elements into standardized output.", "GEN", "Snowflake", "Must Have", "Foundation", "Leadership-aligned segmentation"),
    ("GEN002", "Consolidate standard project details into single-source view.", "GEN", "Snowflake", "Must Have", "Foundation", "Faster activation reporting"),
    ("GEN003", "Ingest technical audit columns for traceability.", "GEN", "Snowflake", "Must Have", "Foundation", "Data integrity and accountability"),
    ("CHL001", "Identify channel activations via Project Type.", "CHL", "Snowflake", "Must Have", "Foundation", "Correct channel categorization"),
    ("CHL002", "Unified Channel Source field from conditional uap channel fields.", "CHL", "Snowflake", "Must Have", "Foundation", "Consistent activation origins"),
    ("CHL003", "Channel Sub-Type for tactical execution type.", "CHL", "Snowflake", "Must Have", "Foundation", "Granular channel insights"),
    ("CHL004", "Capture CCID and DTID for legacy monitoring (informational).", "CHL", "Snowflake", "Should Have", "Foundation", "Legacy adoption monitoring"),
    ("CHL005", "Capture complete activation URL with tracking parameters.", "CHL", "Workfront / Snowflake", "Must Have", "Foundation", "UTM governance audit"),
    ("CON001", "Identify content initiatives via Project Type.", "CON", "Snowflake", "Must Have", "Foundation", "Correct content categorization"),
    ("CON002", "Unified Content Sub Type from conditional content format fields.", "CON", "Snowflake", "Must Have", "Foundation", "Content performance analysis"),
    ("CON003", "Unified Content Platform column from conditional fields.", "CON", "Snowflake", "Must Have", "Foundation", "Track hosting location"),
    ("CON004", "First-Party vs Third-Party classification for events/webinars.", "CON", "Snowflake", "Must Have", "Foundation", "ROI segmentation"),
    ("CON005", "Webinar Event Date and Delivery State for FY27 attribution.", "CON", "Snowflake", "Must Have", "Foundation", "FY27 attribution and follow-up"),
    ("OPS-001", "Daily refresh; Golden Dataset available by 8:00 AM ET.", "Operations", "Data Engineering", "Must Have", "Ongoing", "Current data for daily workflows"),
    ("OPS-002", "Same-day alert on dataset refresh failure.", "Operations", "Data Engineering", "Must Have", "Ongoing", "Prevent stale data consumption"),
])

PROG7 = "CTT & CTT Request Portal FY27 Updates"
DOC7 = "202060730-CTT and CTT Request Portal FY27 Updates.docx"
add_std(PROG7, DOC7, "v1.3 (04/08/2026)", [
    ("CTT001", "CTT bridge fields for Workfront-aligned FY27 attributes.", "CTT Tool", "CTT", "Must Have", "Q1 FY27", "Same FY27 elements during transition"),
    ("CTT002", "Funnel Stage on Activity ID.", "CTT Tool / Activity ID", "CTT", "Must Have", "Q1 FY27", "Funnel position tagging"),
    ("CTT003", "Rename Business Unit to Stakeholder Unit on Activity ID.", "CTT Tool / Activity ID", "CTT", "Must Have", "Q1 FY27", "Consistent stakeholder attribution"),
    ("CTT004", "Technology field with dependent Campaign/Program on Activity ID.", "CTT Tool / Activity ID", "CTT", "Must Have", "Q1 FY27", "Tech > Campaign > Program hierarchy"),
    ("CTT005", "Remove Buying Centre linkage from Activity ID UI.", "CTT Tool / Activity ID", "CTT", "Must Have", "Q1 FY27", "Unconstrained campaign selection"),
    ("CTT006", "GTMPR audit and backend update of historical Activity IDs.", "CTT Tool / Data Migration", "CTT / Dev", "Must Have", "Q1 FY27", "Cross-time reporting accuracy"),
    ("CTT007", "Optional ID Type on Activity ID (Event, Webinar, Driving Channel URL).", "CTT Tool / Activity ID", "CTT", "Should Have", "Q1 FY27", "Execution mechanism classification"),
    ("CTT008", "Funnel Stage on Offer ID.", "CTT Tool / Offer ID", "CTT", "Must Have", "Q1 FY27", "Content funnel stage data"),
    ("CTT009", "Technology/Campaign/Program on Offer ID.", "CTT Tool / Offer ID", "CTT", "Must Have", "Q1 FY27", "Unified hierarchies"),
    ("CTT010", "Stakeholder picklist on Offer ID (was Content Creation Team).", "CTT Tool / Offer ID", "CTT", "Must Have", "Q1 FY27", "Organizational taxonomy alignment"),
    ("CTT011", "Standardized Offer Type by Offer Category.", "CTT Tool / Offer ID", "CTT", "Must Have", "Q1 FY27", "Content categorization"),
    ("CTT012", "Updated Business Entity from Finance master list.", "CTT Tool / Offer ID", "CTT", "Must Have", "Q1 FY27", "Financial governance alignment"),
    ("CTT013", "Primary Offer URL field on Offer ID.", "CTT Tool / Offer ID", "CTT / OMS", "Should Have", "Q1 FY27", "Destination URL for reporting"),
    ("CTT014", "Remove Vertical Market from Offer ID UI; retain OMS schema.", "CTT Tool / Offer ID", "CTT / OMS", "Must Have", "Q1 FY27", "Stop legacy input without breaking pipelines"),
    ("CTT015", "Publish all new CTT field values to OMS tables.", "CTT Tool / OMS", "OMS / Snowflake", "Must Have", "Q1 FY27", "FY27 data for attribution models"),
    ("CTT016", "Exclude new CTT fields from Tray/Allocadia/SFDC APIs during bridge.", "CTT Tool / Integration", "Tray.io / SFDC", "Must Have", "Q1 FY27", "Prevent integration disruption"),
    ("CTT017", "Backups and dry-run before historical backend updates.", "CTT Tool / Data Migration", "Dev Team", "Must Have", "Pre-production", "Rollback capability"),
    ("CTT018", "GTMPR audit and backend update of historical Offer IDs.", "CTT Tool / Data Migration", "CTT / Dev", "Must Have", "Q1 FY27", "Historical Offer alignment"),
    ("POR001", "Portal Stakeholder Group matches CTT Stakeholder Unit (Activity ID).", "CTT Portal / Activity ID", "CTT Request Portal", "Must Have", "Q1 FY27", "Accurate team attribution"),
    ("POR002", "Portal hierarchy Technology > Campaign > Program for Activity IDs.", "CTT Portal / Activity ID", "CTT Request Portal", "Must Have", "Q1 FY27", "Product messaging alignment"),
    ("POR003", "Funnel Stage on Activity ID portal requests.", "CTT Portal / Activity ID", "CTT Request Portal", "Must Have", "Q1 FY27", "Funnel analysis at intake"),
    ("POR004", "Optional ID Type on Activity ID portal workflows.", "CTT Portal / Activity ID", "CTT Request Portal", "Should Have", "Q1 FY27", "Activity type tagging at intake"),
    ("POR005", "Standardized Offer Type on Offer ID portal requests.", "CTT Portal / Offer ID", "CTT Request Portal", "Must Have", "Q1 FY27", "Content format analysis"),
    ("POR006", "Updated Business/Sub Business Entity on Offer ID portal.", "CTT Portal / Offer ID", "CTT Request Portal", "Must Have", "Q1 FY27", "P&L mapping for ROI"),
    ("POR007", "Remove Vertical Market from Offer ID portal forms.", "CTT Portal / Offer ID", "CTT Request Portal", "Must Have", "Q1 FY27", "Reduce intake friction"),
    ("POR008", "Technology > Campaign > Program on Offer ID portal forms.", "CTT Portal / Offer ID", "CTT Request Portal", "Must Have", "Q1 FY27", "Multi-touch attribution linkage"),
    ("POR009", "Funnel Stage on Offer ID portal forms.", "CTT Portal / Offer ID", "CTT Request Portal", "Must Have", "Q1 FY27", "Funnel progression by content"),
    ("POR010", "Primary Offer URL on Offer ID portal pathways.", "CTT Portal / Offer ID", "CTT Request Portal", "Should Have", "Q1 FY27", "Consistent URL capture"),
    ("SFDC001", "No immediate SFDC tracking parameter changes (OCC protection).", "Salesforce", "SFDC / OCC", "Must Have", "Q1 FY27", "No CRM disruption"),
    ("SFDC002", "Existing SFDC campaign alignment to Activity/Offer ID unchanged.", "Salesforce", "SFDC", "Must Have", "Q1 FY27", "Unchanged lead/campaign processes"),
    ("TECH001", "No Eloqua picklist/form/AOJ changes during CTT validation phase.", "Marketing Automation", "Eloqua", "Must Have", "Q1 FY27", "Automation continuity during validation"),
])


def build_workbook(path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Requirements Register"

    ws_summary = wb.create_sheet("Summary by Project", 0)
    ws_summary.append(["Project / BRD", "Total", "Must Have", "Should Have", "Could Have"])
    by_proj = defaultdict(list)
    for r in requirements:
        by_proj[r["Project / BRD"]].append(r)
    for proj, items in sorted(by_proj.items()):
        ws_summary.append([
            proj,
            len(items),
            sum(1 for i in items if i["Priority"] == "Must Have"),
            sum(1 for i in items if i["Priority"] == "Should Have"),
            sum(1 for i in items if i["Priority"] == "Could Have"),
        ])

    ws_stake = wb.create_sheet("Summary by Stakeholder")
    ws_stake.append(["Program / Initiative", "Count"])
    for stake, count in Counter(r["Program / Initiative"] for r in requirements).most_common():
        ws_stake.append([stake, count])

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")

    for col, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx, req in enumerate(requirements, 2):
        for col_idx, h in enumerate(HEADERS, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=req.get(h, ""))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    widths = [14, 60, 28, 22, 22, 18, 12, 16, 40, 16, 30, 25, 25, 22, 20, 45, 18, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(requirements) + 1}"

    ws_inst = wb.create_sheet("How to Use")
    for row in [
        ["Requirements Register - Instructions"],
        [""],
        ["Purpose", "Collates requirements from 7 BRDs to track status, barriers, and impacts."],
        [""],
        ["Complete these columns", "Status | Barriers / Blockers | Impact on Main Project | Owner / Responsible Team | Comments"],
        [""],
        ["Suggested Status values", "Not Started | In Progress | Blocked | In Review | Complete | Deferred | Cancelled"],
        [""],
        ["Total requirements", len(requirements)],
        [""],
        ["Project map", "UIF Phase 1 → Phase 2 → CTT Decommission; CTT FY27 + DTID Alignment are bridge initiatives; Workfront Dataset + POC feed the path"],
    ]:
        ws_inst.append(row)
    ws_inst.column_dimensions["A"].width = 24
    ws_inst.column_dimensions["B"].width = 90

    wb.save(path)
    print(f"Saved {len(requirements)} requirements to {path}")


if __name__ == "__main__":
    build_workbook("/workspace/BRD_Requirements_Register.xlsx")
